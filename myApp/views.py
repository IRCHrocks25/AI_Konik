import csv
import importlib
import io
import json as _json
import logging
import os
import re
import uuid
from datetime import datetime, timedelta

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import check_password, make_password
from django.core.exceptions import ValidationError
from django.core.validators import URLValidator
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncDate
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone as dj_tz
from django.utils.dateparse import parse_datetime as django_parse_datetime
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt

from .admin_audit import log_error, record_admin_action
from .auth_utils import (
    _do_restore_impersonation,
    _is_admin_user,
    admin_required,
    check_impersonation_timeout,
    get_current_user,
    get_request_json,
    login_required_api,
    login_user,
    logout_user,
)
from .models import AdminAction, Agent, AgentPrompt, Banner, ChatMessage, ChatSession, CustomUser, ErrorLog, Event, Industry, Prompt, SavedPrompt, Tool
from .openai_service import (
    MAX_REQUEST_TOKENS,
    estimate_messages_tokens,
    get_openai_reply,
    trim_history_to_budget,
)
from .personalization import build_full_system_prompt
from .seed_data import seed_agents_and_prompts, seed_events, seed_industries, seed_tools

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency runtime check
    load_workbook = None

try:
    import cloudinary
    import cloudinary.uploader
except Exception:  # pragma: no cover - optional dependency runtime check
    cloudinary = None

logger = logging.getLogger(__name__)

SEED_PROMPTS = [
    {
        "title": "Review this NDA for liability exposure",
        "body": "Analyze this NDA and identify liability risks, indemnification issues, and missing protections.",
        "industry": "legal",
        "category": "Contracts",
        "usage_count": 2100,
    },
    {
        "title": "Generate a mutual NDA for software development",
        "body": "Create a mutual NDA for software collaboration with IP protections and a two-year confidentiality period.",
        "industry": "legal",
        "category": "Drafting",
        "usage_count": 1800,
    },
    {
        "title": "Summarize key risks in M&A term sheet",
        "body": "Review this term sheet and summarize top seller-side risks and unusual clauses.",
        "industry": "legal",
        "category": "M&A",
        "usage_count": 1400,
    },
]

INDUSTRY_MAP = {
    "legal": "legal",
    "healthcare": "healthcare",
    "finance": "finance",
    "marketing": "marketing",
    "technology": "technology",
    "real estate": "realestate",
    "realestate": "realestate",
    "accounting": "accounting",
    "logistics": "logistics",
    "manufacturing": "manufacturing",
}

_PROFILE_CHOICES = {
    "company_size": {"solo", "2-10", "11-50", "51-200", "200+"},
    "years_experience": {"<1", "1-3", "3-7", "7-15", "15+"},
    "communication_style": {"direct", "friendly", "formal", "casual"},
    "response_length": {"concise", "balanced", "detailed"},
    "expertise_level": {"beginner", "intermediate", "expert"},
    "emoji_use": {"never", "sparingly", "freely"},
    "pushback_style": {"always", "diplomatic", "supportive"},
    "explanation_style": {"answer_only", "with_reasoning", "step_by_step"},
    "clarifying_questions": {"when_needed", "always", "never"},
}

_PROFILE_TEXT_LIMITS = {
    "display_name": 80,
    "role": 120,
    "timezone": 64,
    "current_focus": 500,
    "things_to_avoid": 300,
    "about_me": 1000,
}

_EXPERTISE_AREAS_VALID = {
    "Legal", "Finance", "Marketing", "Tech", "Healthcare", "RealEstate",
    "Accounting", "Logistics", "Manufacturing", "Operations", "Sales",
    "HR", "Strategy", "Product", "Design",
}

HEADER_ALIASES = {
    "id": "id",
    "industry": "industry",
    "category": "category",
    "subcategory": "sub_category",
    "subcategory": "sub_category",
    "sub-category": "sub_category",
    "prompttitle": "prompt_title",
    "prompt title": "prompt_title",
    "promptfulltext": "prompt_full_text",
    "prompt(fulltext)": "prompt_full_text",
    "prompt (full text)": "prompt_full_text",
    "difficulty": "difficulty",
    "usecasetype": "use_case_type",
    "use case type": "use_case_type",
    "estimatedtimesaved": "estimated_time_saved",
    "estimated time saved": "estimated_time_saved",
    "tags": "tags",
}

LEGACY_PATH_REDIRECTS = {
    "index.html": "/index/",
    "dashboard.html": "/dashboard/",
    "agents.html": "/agents/",
    "agent-chat.html": "/agent-chat/",
    "prompts.html": "/prompts/",
    "industries.html": "/industries/",
    "events.html": "/events/",
    "tools.html": "/tools/",
    "consulting.html": "/consulting/",
    "billing.html": "/billing/",
    "profile.html": "/profile/",
    "settings.html": "/settings/",
    "login.html": "/login/",
    "register.html": "/register/",
    "prompt-import.html": "/prompt-import/",
}


def ensure_seed_prompts():
    if Prompt.objects.exists():
        return
    Prompt.objects.bulk_create([Prompt(**prompt) for prompt in SEED_PROMPTS])


def ensure_seed_agents():
    seed_agents_and_prompts()


def ensure_seed_industries():
    if Industry.objects.exists():
        return
    seed_industries()


def ensure_seed_events():
    if Event.objects.exists():
        return
    ensure_seed_industries()  # events reference industries
    seed_events()


def ensure_seed_tools():
    if Tool.objects.exists():
        return
    seed_tools()


def _parse_event_date(raw):
    """Parse an ISO-8601 string into a tz-aware datetime, or return None."""
    val = str(raw or "").strip()
    if not val:
        return None
    dt = django_parse_datetime(val)
    if dt is not None:
        return dj_tz.make_aware(dt) if dt.tzinfo is None else dt
    from datetime import datetime as _dt
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return dj_tz.make_aware(_dt.strptime(val, fmt))
        except ValueError:
            continue
    return None


def home(request):
    return render(request, "index.html")


def dashboard(request):
    return render(request, "dashboard.html")


def agents(request):
    return render(request, "agents.html")


def agent_chat(request):
    return render(request, "agent-chat.html")


def prompts(request):
    return render(request, "prompts.html")


def industries(request):
    return render(request, "industries.html")


def events(request):
    return render(request, "events.html")


def tools(request):
    return render(request, "tools.html")


def consulting(request):
    return render(request, "consulting.html")


def billing(request):
    return render(request, "billing.html")


def profile_page(request):
    return render(request, "profile.html")


def settings_page(request):
    return render(request, "settings.html")


def login(request):
    return render(request, "login.html")


def register(request):
    return render(request, "register.html")


def shared_css(request):
    return render(request, "shared.css", content_type="text/css")


def legacy_html_redirect(request, page):
    target = LEGACY_PATH_REDIRECTS.get(page)
    if not target:
        raise Http404("Legacy path not found")
    query = request.META.get("QUERY_STRING", "")
    if query:
        return redirect(f"{target}?{query}")
    return redirect(target)


@admin_required
def admin_dashboard(request, **kwargs):
    return render(request, "admin-dashboard.html")


def prompt_import_dashboard(request):
    user = get_current_user(request)
    if not user:
        return redirect("login")
    if not _is_admin_user(user):
        return redirect("prompts")
    return render(request, "prompt-import.html")


@csrf_exempt
def api_register(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)

    email = str(payload.get("email", "")).strip().lower()
    password = str(payload.get("password", "")).strip()
    first_name = str(payload.get("first_name", "")).strip()
    last_name = str(payload.get("last_name", "")).strip()
    industry = str(payload.get("industry", "")).strip()

    if not email or not password or not first_name:
        return JsonResponse(
            {"error": "first_name, email, and password are required"}, status=400
        )
    if CustomUser.objects.filter(email=email).exists():
        return JsonResponse({"error": "Email already registered"}, status=409)

    user = CustomUser.objects.create(
        first_name=first_name,
        last_name=last_name,
        email=email,
        industry=industry,
        password_hash=make_password(password),
    )
    login_user(request, user)
    return JsonResponse(
        {
            "id": user.id,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "industry": user.industry,
        },
        status=201,
    )


@csrf_exempt
def api_login(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)

    email = str(payload.get("email", "")).strip().lower()
    password = str(payload.get("password", "")).strip()
    user = CustomUser.objects.filter(email=email).first()
    if not user or not check_password(password, user.password_hash):
        # Bridge Django auth users (including superusers) into CustomUser login.
        django_user_model = get_user_model()
        django_user = django_user_model.objects.filter(
            Q(email__iexact=email) | Q(username__iexact=email)
        ).first()
        if not django_user or not django_user.check_password(password):
            return JsonResponse({"error": "Invalid email or password"}, status=401)

        bridge_email = (django_user.email or "").strip().lower()
        if not bridge_email:
            # Ensure a valid email for CustomUser unique field.
            bridge_email = f"{django_user.username}@local.user"

        user, _ = CustomUser.objects.get_or_create(
            email=bridge_email,
            defaults={
                "first_name": django_user.first_name or django_user.username or "User",
                "last_name": django_user.last_name or "",
                "industry": "",
                "password_hash": django_user.password,
            },
        )
        # Keep the bridge record synced to Django auth credentials.
        fields_to_update = []
        desired_first = django_user.first_name or django_user.username or "User"
        desired_last = django_user.last_name or ""
        if user.first_name != desired_first:
            user.first_name = desired_first
            fields_to_update.append("first_name")
        if user.last_name != desired_last:
            user.last_name = desired_last
            fields_to_update.append("last_name")
        if user.password_hash != django_user.password:
            user.password_hash = django_user.password
            fields_to_update.append("password_hash")
        if fields_to_update:
            user.save(update_fields=fields_to_update)

    login_user(request, user)
    return JsonResponse(
        {
            "id": user.id,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "industry": user.industry,
        }
    )


@csrf_exempt
def api_logout(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    logout_user(request)
    return JsonResponse({"success": True})


def api_me(request):
    check_impersonation_timeout(request)  # must run before get_current_user reads the session
    user = get_current_user(request)
    if not user:
        return JsonResponse({"authenticated": False, "is_admin": False})
    is_impersonating = "_impersonated_by" in request.session
    result = {
        "authenticated": True,
        "is_admin": _is_admin_user(user),
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "industry": user.industry,
        "is_impersonating": is_impersonating,
    }
    if is_impersonating:
        admin_id = request.session.get("_impersonated_by")
        admin_custom = CustomUser.objects.filter(id=admin_id).first()
        result["impersonated_by_email"] = admin_custom.email if admin_custom else None
    return JsonResponse(result)


def _validate_and_build_profile_updates(payload):
    """Validate a profile payload and return (updates, errors).

    updates: dict of field → validated value, ready to setattr on a CustomUser.
    errors:  dict of field → error message for invalid inputs.

    Handles all 18 personalization fields. is_suspended is intentionally NOT
    handled here — callers that need it validate and append it themselves so
    that admin-specific business rules (cannot suspend self/other-admin) stay
    in the view layer, not in this shared helper.
    """
    errors, updates = {}, {}

    for field, limit in _PROFILE_TEXT_LIMITS.items():
        if field not in payload:
            continue
        val = str(payload[field]).strip()
        if len(val) > limit:
            errors[field] = f"Too long (max {limit} characters)."
        else:
            updates[field] = val

    for field, valid_set in _PROFILE_CHOICES.items():
        if field not in payload:
            continue
        val = str(payload[field]).strip()
        if val and val not in valid_set:
            errors[field] = f"Invalid choice. Allowed: {', '.join(sorted(valid_set))}."
        else:
            updates[field] = val

    if "industry" in payload:
        updates["industry"] = _normalize_industry(str(payload.get("industry", "")))

    if "formality" in payload:
        try:
            v = int(payload["formality"])
            if not 1 <= v <= 5:
                raise ValueError
            updates["formality"] = v
        except (TypeError, ValueError):
            errors["formality"] = "Must be an integer between 1 and 5."

    if "expertise_areas" in payload:
        val = payload["expertise_areas"]
        if not isinstance(val, list):
            errors["expertise_areas"] = "Must be a list."
        else:
            invalid = [v for v in val if v not in _EXPERTISE_AREAS_VALID]
            if invalid:
                errors["expertise_areas"] = f"Invalid values: {', '.join(sorted(invalid))}."
            else:
                updates["expertise_areas"] = val

    return updates, errors


@csrf_exempt
@login_required_api
def api_profile(request):
    user = request.current_user

    if request.method == "GET":
        return JsonResponse({
            "display_name": user.display_name,
            "role": user.role,
            "industry": user.industry,
            "company_size": user.company_size,
            "years_experience": user.years_experience,
            "timezone": user.timezone,
            "communication_style": user.communication_style,
            "response_length": user.response_length,
            "expertise_level": user.expertise_level,
            "formality": user.formality,
            "emoji_use": user.emoji_use,
            "pushback_style": user.pushback_style,
            "explanation_style": user.explanation_style,
            "clarifying_questions": user.clarifying_questions,
            "expertise_areas": user.expertise_areas,
            "current_focus": user.current_focus,
            "things_to_avoid": user.things_to_avoid,
            "about_me": user.about_me,
        })

    if request.method != "PUT":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)

    updates, errors = _validate_and_build_profile_updates(payload)
    if errors:
        return JsonResponse({"errors": errors}, status=400)

    if not updates:
        return JsonResponse({"success": True})

    fields_to_save = list(updates.keys())

    if user.profile_completed_at is None:
        has_meaningful = any(
            (isinstance(v, list) and v) or (isinstance(v, str) and v)
            for v in updates.values()
        )
        if has_meaningful:
            user.profile_completed_at = dj_tz.now()
            fields_to_save.append("profile_completed_at")

    for field, value in updates.items():
        setattr(user, field, value)

    fields_to_save.append("updated_at")
    user.save(update_fields=fields_to_save)
    return JsonResponse({"success": True})


_HEX_COLOR_RE = re.compile(r"^#(?:[0-9a-fA-F]{3,4}|[0-9a-fA-F]{6}|[0-9a-fA-F]{8})$")


def _is_valid_hex_color(value):
    """True iff value is a hex color: #RGB, #RGBA, #RRGGBB, or #RRGGBBAA.

    Hex-only validation defends against stored XSS via the accent_bg field —
    rejecting anything that could end up in a style attribute or CSS custom
    property as raw user input.
    """
    return bool(_HEX_COLOR_RE.match(str(value or "").strip()))


def _serialize_agent(agent, include_prompts=False):
    data = {
        "id": agent.id,
        "name": agent.name,
        "slug": agent.slug,
        "industry": agent.industry,
        "description": agent.description,
        "icon_class": agent.icon_class,
        "accent_bg": agent.accent_bg,
        "tag": agent.tag,
        "usage_count": agent.usage_count,
        "is_featured": agent.is_featured,
        "is_active": agent.is_active,
        "sort_order": agent.sort_order,
    }
    if include_prompts:
        prompts = list(agent.prompts.all().values("id", "prompt_type", "content", "sort_order"))
        data["hints"] = [p["content"] for p in prompts if p["prompt_type"] == "hint"]
        data["use_cases"] = [p["content"] for p in prompts if p["prompt_type"] == "use_case"]
        data["prompts"] = prompts
    return data


def _to_non_negative_int(value, default=0):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(parsed, 0)


def api_agents(request):
    ensure_seed_agents()
    include_prompts = request.GET.get("include_prompts", "").strip() == "1"
    industry = request.GET.get("industry", "").strip().lower()
    queryset = Agent.objects.filter(is_active=True)
    if industry and industry != "all":
        queryset = queryset.filter(industry__iexact=industry)
    return JsonResponse(
        {
            "items": [
                _serialize_agent(agent, include_prompts=include_prompts)
                for agent in queryset.prefetch_related("prompts")
            ]
        }
    )


@csrf_exempt
@login_required_api
def api_admin_agents(request):
    ensure_seed_agents()
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    if request.method == "GET":
        queryset = Agent.objects.all().prefetch_related("prompts")
        return JsonResponse({"items": [_serialize_agent(agent, include_prompts=True) for agent in queryset]})

    if request.method == "POST":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)
        name = str(payload.get("name", "")).strip()
        industry = str(payload.get("industry", "")).strip().lower()
        description = str(payload.get("description", "")).strip()
        errors = {}
        if not name:
            errors['name'] = 'Required.'
        if not industry:
            errors['industry'] = 'Required.'
        if not description:
            errors['description'] = 'Required.'
        if errors:
            return JsonResponse({'errors': errors}, status=400)
        slug = slugify(name)
        if not slug:
            return JsonResponse({"error": "Invalid name"}, status=400)
        if Agent.objects.filter(Q(name__iexact=name) | Q(slug=slug)).exists():
            return JsonResponse({"error": "Agent already exists"}, status=409)

        accent_bg = str(payload.get("accent_bg", "#EEF2FF")).strip() or "#EEF2FF"
        if not _is_valid_hex_color(accent_bg):
            return JsonResponse({"error": "accent_bg must be a hex color like #RRGGBB"}, status=400)

        # hints / use_cases are optional on POST — the separate /api/admin/agents/<id>/prompts
        # endpoint also creates AgentPrompts and is preserved for backwards compatibility.
        # The admin-dashboard flow supplies both arrays here so creation is one round-trip.
        raw_hints = payload.get("hints") or []
        raw_use_cases = payload.get("use_cases") or []
        if not isinstance(raw_hints, list) or not isinstance(raw_use_cases, list):
            return JsonResponse({"error": "hints and use_cases must be arrays"}, status=400)
        hints = [str(h).strip() for h in raw_hints if str(h).strip()][:8]
        use_cases = [str(u).strip() for u in raw_use_cases if str(u).strip()][:8]

        with transaction.atomic():
            agent = Agent.objects.create(
                name=name,
                slug=slug,
                industry=industry,
                description=description,
                icon_class=str(payload.get("icon_class", "fa-robot")).strip() or "fa-robot",
                accent_bg=accent_bg,
                tag=str(payload.get("tag", "")).strip(),
                usage_count=_to_non_negative_int(payload.get("usage_count", 0), default=0),
                is_featured=bool(payload.get("is_featured", False)),
                is_active=bool(payload.get("is_active", True)),
                sort_order=_to_non_negative_int(payload.get("sort_order", 100), default=100),
            )
            AgentPrompt.objects.bulk_create(
                [
                    AgentPrompt(agent=agent, prompt_type="hint", content=h, sort_order=(i + 1) * 10)
                    for i, h in enumerate(hints)
                ]
                + [
                    AgentPrompt(agent=agent, prompt_type="use_case", content=u, sort_order=(i + 1) * 10)
                    for i, u in enumerate(use_cases)
                ]
            )

        record_admin_action(
            admin=request.current_user,
            action_type="agent.create",
            target_type="agent",
            target_id=agent.id,
            metadata={
                "name": agent.name,
                "industry": agent.industry,
                "hints_count": len(hints),
                "use_cases_count": len(use_cases),
            },
        )
        return JsonResponse(_serialize_agent(agent, include_prompts=True), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# Field → max-length cap for PATCH text updates. industry/description checked
# at save time the same way as POST. Caps are generous; UI enforces tighter
# practical limits.
_AGENT_PATCH_TEXT_FIELDS = {
    "name": 160,
    "industry": 100,
    "description": 4000,
    "tag": 120,
    "icon_class": 80,
}


@csrf_exempt
@login_required_api
def api_admin_agent_detail(request, agent_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    agent = Agent.objects.filter(id=agent_id).first()
    if not agent:
        return JsonResponse({"error": "Agent not found"}, status=404)

    if request.method == "DELETE":
        agent_name = agent.name
        prompts_count = agent.prompts.count()
        agent.delete()  # FK CASCADE removes AgentPrompt rows
        record_admin_action(
            admin=request.current_user,
            action_type="agent.delete",
            target_type="agent",
            target_id=agent_id,
            metadata={"name": agent_name, "prompts_deleted": prompts_count},
        )
        return JsonResponse({"deleted": True})

    if request.method == "PATCH":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        updates = {}

        for field, limit in _AGENT_PATCH_TEXT_FIELDS.items():
            if field not in payload:
                continue
            val = str(payload[field]).strip()
            if field in ("name", "industry", "description") and not val:
                errors[field] = "Cannot be empty."
                continue
            if len(val) > limit:
                errors[field] = f"Too long (max {limit})."
                continue
            if field == "industry":
                val = val.lower()
            updates[field] = val

        # Re-slug when name changes; reject collisions with other agents.
        if "name" in updates and updates["name"] != agent.name:
            new_slug = slugify(updates["name"])
            if not new_slug:
                errors["name"] = "Invalid name."
            elif Agent.objects.filter(
                Q(name__iexact=updates["name"]) | Q(slug=new_slug)
            ).exclude(id=agent.id).exists():
                errors["name"] = "Another agent with this name already exists."
            else:
                updates["slug"] = new_slug

        if "accent_bg" in payload:
            val = str(payload["accent_bg"]).strip()
            if not _is_valid_hex_color(val):
                errors["accent_bg"] = "Must be a hex color like #RRGGBB."
            else:
                updates["accent_bg"] = val

        for bool_field in ("is_featured", "is_active"):
            if bool_field in payload:
                if not isinstance(payload[bool_field], bool):
                    errors[bool_field] = "Must be a boolean."
                else:
                    updates[bool_field] = payload[bool_field]

        for int_field in ("usage_count", "sort_order"):
            if int_field in payload:
                updates[int_field] = _to_non_negative_int(
                    payload[int_field], default=getattr(agent, int_field)
                )

        replace_hints = "hints" in payload
        replace_use_cases = "use_cases" in payload
        new_hints = []
        new_use_cases = []
        if replace_hints:
            if not isinstance(payload["hints"], list):
                errors["hints"] = "Must be an array."
            else:
                new_hints = [str(h).strip() for h in payload["hints"] if str(h).strip()][:8]
        if replace_use_cases:
            if not isinstance(payload["use_cases"], list):
                errors["use_cases"] = "Must be an array."
            else:
                new_use_cases = [str(u).strip() for u in payload["use_cases"] if str(u).strip()][:8]

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        changed_fields = list(updates.keys())
        with transaction.atomic():
            if updates:
                for field, value in updates.items():
                    setattr(agent, field, value)
                agent.save(update_fields=changed_fields + ["updated_at"])
            if replace_hints:
                agent.prompts.filter(prompt_type="hint").delete()
                AgentPrompt.objects.bulk_create([
                    AgentPrompt(agent=agent, prompt_type="hint", content=h, sort_order=(i + 1) * 10)
                    for i, h in enumerate(new_hints)
                ])
                changed_fields.append("hints")
            if replace_use_cases:
                agent.prompts.filter(prompt_type="use_case").delete()
                AgentPrompt.objects.bulk_create([
                    AgentPrompt(agent=agent, prompt_type="use_case", content=u, sort_order=(i + 1) * 10)
                    for i, u in enumerate(new_use_cases)
                ])
                changed_fields.append("use_cases")

        record_admin_action(
            admin=request.current_user,
            action_type="agent.update",
            target_type="agent",
            target_id=agent.id,
            metadata={"changed_fields": changed_fields},
        )
        agent.refresh_from_db()
        return JsonResponse(_serialize_agent(agent, include_prompts=True))

    return JsonResponse({"error": "Method not allowed"}, status=405)


# Fixed prompt template for agent AI generation. Single user message — no
# system prompt — because this is one-shot structured generation, not a chat.
# Curly braces around the JSON example are doubled so str.format leaves them
# untouched. Token cost: ~250 prompt + ~400 response ≈ 650 per call.
_AGENT_GENERATION_PROMPT_TEMPLATE = """You are configuring a new AI agent for AI KONIK, a business assistant platform.

Agent details:
- Name: {name}
- Industry: {industry}
- Description: {description}
- Tag: {tag}

Generate suggestions in valid JSON. Strict requirements:

1. "hints" — array of EXACTLY 4 short example user queries (5-12 words each, imperative voice). These appear as quick-start buttons in the chat UI.

2. "use_cases" — array of EXACTLY 4 longer example requests (15-30 words each, more detailed than hints). These appear as previewable examples.

3. "accent_bg" — exactly one hex color in #RRGGBB format. Choose a soft pastel that fits the industry's emotional register:
   - legal/finance → cool blues like #EEF2FF
   - healthcare → gentle greens like #ECFDF5
   - marketing/design → warm pinks like #FFF1F2
   - technology → light blues like #F0F9FF
   - real estate → light violets like #F5F3FF
   - accounting → mint greens like #ECFDF5
   - logistics/manufacturing → warm neutrals like #FFF7ED or #F8FAFC
   - other → fall back to a neutral cool tone like #F1F5F9

Return ONLY this JSON structure, with NO markdown fences and NO explanation:

{{
  "hints": ["...", "...", "...", "..."],
  "use_cases": ["...", "...", "...", "..."],
  "accent_bg": "#XXXXXX"
}}"""


@csrf_exempt
@login_required_api
def api_admin_agent_generate(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)

    name = str(payload.get("name", "")).strip()
    industry = str(payload.get("industry", "")).strip()
    description = str(payload.get("description", "")).strip()
    tag = str(payload.get("tag", "")).strip()

    if not name or not industry or not description:
        return JsonResponse(
            {"error": "name, industry, and description are required"}, status=400
        )
    if len(name) > 200:
        return JsonResponse({"error": "name must be ≤200 characters"}, status=400)
    if len(description) > 1000:
        return JsonResponse({"error": "description must be ≤1000 characters"}, status=400)

    prompt_text = _AGENT_GENERATION_PROMPT_TEMPLATE.format(
        name=name, industry=industry, description=description, tag=tag or "(none)",
    )

    # Audit the attempt before calling OpenAI so failures are still recorded.
    record_admin_action(
        admin=request.current_user,
        action_type="agent.generate",
        target_type="agent",
        metadata={"name_input": name, "industry_input": industry},
    )

    try:
        raw_text, _ = get_openai_reply([{"role": "user", "content": prompt_text}])
    except Exception as exc:
        logger.exception("Agent generate failed: %s", exc)
        log_error(
            error_type="openai_agent_generate",
            message=str(exc),
            user=request.current_user,
            metadata={"name_input": name},
        )
        return JsonResponse({"error": "Generation failed. Please try again."}, status=502)

    cleaned = (raw_text or "").strip()
    # Defensive: strip ```json fences if model added them despite instructions.
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        data = _json.loads(cleaned)
    except (ValueError, TypeError):
        return JsonResponse(
            {"error": "Generation produced invalid output, please retry."},
            status=502,
        )

    warnings = []
    hints_out = []
    use_cases_out = []
    accent_out = ""

    raw_hints = data.get("hints") if isinstance(data, dict) else None
    if (
        isinstance(raw_hints, list)
        and len(raw_hints) == 4
        and all(isinstance(h, str) and h.strip() for h in raw_hints)
    ):
        hints_out = [h.strip() for h in raw_hints]
    else:
        warnings.append("hints")

    raw_use_cases = data.get("use_cases") if isinstance(data, dict) else None
    if (
        isinstance(raw_use_cases, list)
        and len(raw_use_cases) == 4
        and all(isinstance(u, str) and u.strip() for u in raw_use_cases)
    ):
        use_cases_out = [u.strip() for u in raw_use_cases]
    else:
        warnings.append("use_cases")

    raw_accent = data.get("accent_bg", "") if isinstance(data, dict) else ""
    if isinstance(raw_accent, str) and _is_valid_hex_color(raw_accent.strip()):
        accent_out = raw_accent.strip()
    else:
        warnings.append("accent_bg")

    # If every field failed validation the model output is fundamentally broken;
    # return 502 so admin retries instead of starting from a blank slate.
    if len(warnings) == 3:
        return JsonResponse(
            {"error": "Generation produced unusable output, please retry."},
            status=502,
        )

    return JsonResponse({
        "hints": hints_out,
        "use_cases": use_cases_out,
        "accent_bg": accent_out,
        "warnings": warnings,
    })


@csrf_exempt
@login_required_api
def api_admin_agent_prompts(request, agent_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    agent = Agent.objects.filter(id=agent_id).first()
    if not agent:
        return JsonResponse({"error": "Agent not found"}, status=404)
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)
    prompt_type = str(payload.get("prompt_type", "hint")).strip().lower()
    content = str(payload.get("content", "")).strip()
    sort_order = _to_non_negative_int(payload.get("sort_order", 100), default=100)
    if prompt_type not in {"hint", "use_case"}:
        return JsonResponse({"error": "prompt_type must be hint or use_case"}, status=400)
    if not content:
        return JsonResponse({"error": "content is required"}, status=400)
    prompt = AgentPrompt.objects.create(
        agent=agent,
        prompt_type=prompt_type,
        content=content,
        sort_order=sort_order,
    )
    return JsonResponse(
        {
            "id": prompt.id,
            "agent_id": agent.id,
            "prompt_type": prompt.prompt_type,
            "content": prompt.content,
            "sort_order": prompt.sort_order,
        },
        status=201,
    )


@csrf_exempt
@login_required_api
def api_admin_agent_prompt_detail(request, prompt_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    prompt = AgentPrompt.objects.filter(id=prompt_id).first()
    if not prompt:
        return JsonResponse({"error": "Prompt not found"}, status=404)
    if request.method == "DELETE":
        prompt.delete()
        return JsonResponse({"deleted": True})
    return JsonResponse({"error": "Method not allowed"}, status=405)


def _normalize_header(header):
    cleaned = re.sub(r"[^a-z0-9]+", "", str(header or "").strip().lower())
    return HEADER_ALIASES.get(cleaned, cleaned)


def _normalize_industry(raw_industry):
    key = str(raw_industry or "").strip().lower()
    return INDUSTRY_MAP.get(key, key)


def _extract_rows_from_upload(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    rows = []
    if ext == ".csv":
        content = uploaded_file.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(content))
        for raw_row in reader:
            normalized = {_normalize_header(k): (v or "").strip() for k, v in raw_row.items()}
            rows.append(normalized)
        return rows
    if ext in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise ValueError("openpyxl is required for .xlsx uploads")
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        headers = next(row_iter, None)
        if not headers:
            return []
        normalized_headers = [_normalize_header(h) for h in headers]
        for raw_values in row_iter:
            if not raw_values or all(v is None or str(v).strip() == "" for v in raw_values):
                continue
            row = {}
            for idx, value in enumerate(raw_values):
                key = normalized_headers[idx] if idx < len(normalized_headers) else f"col_{idx}"
                row[key] = str(value).strip() if value is not None else ""
            rows.append(row)
        return rows
    raise ValueError("Unsupported file type. Use .csv or .xlsx")


def _extract_text_from_uploaded_file(file_name, file_bytes):
    ext = os.path.splitext(file_name)[1].lower()
    text = ""
    if ext in {".txt", ".md", ".csv", ".json"}:
        text = file_bytes.decode("utf-8", errors="ignore")
    elif ext == ".pdf":
        try:
            pypdf_module = importlib.import_module("pypdf")
            PdfReader = getattr(pypdf_module, "PdfReader")
            reader = PdfReader(io.BytesIO(file_bytes))
            parts = []
            for page in reader.pages[:20]:
                extracted = page.extract_text() or ""
                if extracted.strip():
                    parts.append(extracted)
            text = "\n".join(parts)
        except Exception:
            text = ""
    return text.strip()


def _upload_to_cloudinary(file_name, file_bytes):
    if not (
        cloudinary
        and settings.CLOUDINARY_CLOUD_NAME
        and settings.CLOUDINARY_API_KEY
        and settings.CLOUDINARY_API_SECRET
    ):
        return None
    cloudinary.config(
        cloud_name=settings.CLOUDINARY_CLOUD_NAME,
        api_key=settings.CLOUDINARY_API_KEY,
        api_secret=settings.CLOUDINARY_API_SECRET,
        secure=True,
    )
    uploaded = cloudinary.uploader.upload(
        io.BytesIO(file_bytes),
        resource_type="raw",
        folder="aikonik/chat_uploads",
        public_id=f"{uuid.uuid4()}-{os.path.splitext(file_name)[0][:40]}",
        use_filename=False,
        overwrite=True,
    )
    return uploaded.get("secure_url")


@csrf_exempt
@login_required_api
def api_admin_import_prompts(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    try:
        rows = _extract_rows_from_upload(uploaded_file)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    required_keys = {"industry", "category", "prompt_title", "prompt_full_text"}
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        if not required_keys.issubset(set(row.keys())):
            missing = sorted(list(required_keys.difference(set(row.keys()))))
            return JsonResponse(
                {"error": f"Missing required columns: {', '.join(missing)}"}, status=400
            )

        title = row.get("prompt_title", "").strip()
        body = row.get("prompt_full_text", "").strip()
        industry = _normalize_industry(row.get("industry", ""))
        category = row.get("category", "").strip()
        sub_category = row.get("sub_category", "").strip()
        tags = row.get("tags", "").strip()
        difficulty = row.get("difficulty", "").strip()
        use_case_type = row.get("use_case_type", "").strip()
        estimated_time_saved = row.get("estimated_time_saved", "").strip()

        if not title or not body or not industry:
            skipped += 1
            errors.append({"row": index, "error": "Missing title/body/industry"})
            continue

        normalized_category = category or "General"
        if sub_category:
            normalized_category = f"{normalized_category} / {sub_category}"
        metadata_parts = [
            f"Difficulty: {difficulty}" if difficulty else "",
            f"Use Case Type: {use_case_type}" if use_case_type else "",
            f"Estimated Time Saved: {estimated_time_saved}" if estimated_time_saved else "",
            f"Tags: {tags}" if tags else "",
        ]
        metadata = " | ".join([part for part in metadata_parts if part])
        full_body = body if not metadata else f"{body}\n\n[{metadata}]"

        prompt, was_created = Prompt.objects.get_or_create(
            title=title,
            industry=industry,
            defaults={
                "body": full_body,
                "category": normalized_category,
                "status": "approved",
                "created_by": request.current_user,
            },
        )
        if was_created:
            created += 1
            continue

        changed = False
        if prompt.body != full_body:
            prompt.body = full_body
            changed = True
        if prompt.category != normalized_category:
            prompt.category = normalized_category
            changed = True
        if changed:
            prompt.save(update_fields=["body", "category"])
            updated += 1
        else:
            skipped += 1

    return JsonResponse(
        {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:50],
        }
    )


def _serialize_prompt(prompt):
    return {
        "id": prompt.id,
        "title": prompt.title,
        "body": prompt.body,
        "industry": prompt.industry,
        "category": prompt.category,
        "status": prompt.status,
        "usage_count": prompt.usage_count,
        "saved_count": getattr(prompt, "saved_count", 0),
        "created_by_id": prompt.created_by_id,
        "created_at": prompt.created_at.isoformat(),
    }


_PROMPT_STATUS_VALUES = {"approved", "pending"}

_PROMPT_PATCH_TEXT_FIELDS = {
    "title": 255,
    "body": 50000,
    "industry": 100,
    "category": 100,
}


@csrf_exempt
@login_required_api
def api_admin_prompts(request):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    if request.method == "GET":
        try:
            page = max(1, int(request.GET.get("page", 1) or 1))
        except (ValueError, TypeError):
            page = 1
        try:
            page_size = min(200, max(1, int(request.GET.get("page_size", 50) or 50)))
        except (ValueError, TypeError):
            page_size = 50

        search = request.GET.get("search", "").strip()
        industry = request.GET.get("industry", "").strip().lower()
        status_filter = request.GET.get("status", "").strip()
        sort = request.GET.get("sort", "created_desc").strip()

        qs = Prompt.objects.annotate(saved_count=Count("saved_by"))

        if search:
            qs = qs.filter(
                Q(title__icontains=search)
                | Q(body__icontains=search)
                | Q(category__icontains=search)
            )
        if industry:
            qs = qs.filter(industry__iexact=INDUSTRY_MAP.get(industry, industry))
        if status_filter:
            qs = qs.filter(status=status_filter)

        sort_map = {
            "created_desc": "-created_at",
            "created_asc": "created_at",
            "title_asc": "title",
            "saved_count_desc": "-saved_count",
        }
        qs = qs.order_by(sort_map.get(sort, "-created_at"))

        total = qs.count()
        pages = max(1, (total + page_size - 1) // page_size)
        page = min(page, pages)
        offset = (page - 1) * page_size

        return JsonResponse({
            "items": [_serialize_prompt(p) for p in qs[offset: offset + page_size]],
            "page": page,
            "page_size": page_size,
            "total": total,
            "pages": pages,
        })

    if request.method == "POST":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        title = str(payload.get("title", "")).strip()
        body = str(payload.get("body", "")).strip()
        industry = str(payload.get("industry", "")).strip()
        category = str(payload.get("category", "")).strip() or "General"
        status_val = str(payload.get("status", "approved")).strip()

        errors = {}
        if not title:
            errors["title"] = "Required."
        elif len(title) > 255:
            errors["title"] = "Too long (max 255)."
        if not body:
            errors["body"] = "Required."
        elif len(body) > 50000:
            errors["body"] = "Too long (max 50,000)."
        if not industry:
            errors["industry"] = "Required."
        if len(category) > 100:
            errors["category"] = "Too long (max 100)."
        if status_val not in _PROMPT_STATUS_VALUES:
            errors["status"] = "Must be 'approved' or 'pending'."
        if errors:
            return JsonResponse({"errors": errors}, status=400)

        prompt = Prompt.objects.create(
            title=title,
            body=body,
            industry=_normalize_industry(industry),
            category=category,
            status=status_val,
            created_by=request.current_user,
        )
        record_admin_action(
            admin=request.current_user,
            action_type="prompt.create",
            target_type="prompt",
            target_id=prompt.id,
            metadata={"title": prompt.title, "industry": prompt.industry},
        )
        prompt.saved_count = 0
        return JsonResponse(_serialize_prompt(prompt), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required_api
def api_admin_prompt_detail(request, prompt_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    prompt = Prompt.objects.filter(id=prompt_id).annotate(saved_count=Count("saved_by")).first()
    if not prompt:
        return JsonResponse({"error": "Prompt not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(_serialize_prompt(prompt))

    if request.method == "DELETE":
        prompt_title = prompt.title
        prompt.delete()
        record_admin_action(
            admin=request.current_user,
            action_type="prompt.delete",
            target_type="prompt",
            target_id=prompt_id,
            metadata={"title": prompt_title},
        )
        return JsonResponse({"deleted": True})

    if request.method == "PATCH":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        updates = {}

        for field, limit in _PROMPT_PATCH_TEXT_FIELDS.items():
            if field not in payload:
                continue
            val = str(payload[field]).strip()
            if field in ("title", "body", "industry") and not val:
                errors[field] = "Cannot be empty."
                continue
            if len(val) > limit:
                errors[field] = f"Too long (max {limit})."
                continue
            if field == "industry":
                val = _normalize_industry(val)
            updates[field] = val

        if "status" in payload:
            val = str(payload["status"]).strip()
            if val not in _PROMPT_STATUS_VALUES:
                errors["status"] = "Must be 'approved' or 'pending'."
            else:
                updates["status"] = val

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        changed_fields = list(updates.keys())
        if updates:
            for field, value in updates.items():
                setattr(prompt, field, value)
            prompt.save(update_fields=changed_fields)

        record_admin_action(
            admin=request.current_user,
            action_type="prompt.update",
            target_type="prompt",
            target_id=prompt.id,
            metadata={"changed_fields": changed_fields},
        )
        prompt = Prompt.objects.filter(id=prompt_id).annotate(saved_count=Count("saved_by")).first()
        return JsonResponse(_serialize_prompt(prompt))

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required_api
def api_admin_prompts_categories(request):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    categories = list(
        Prompt.objects.exclude(category="")
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )
    return JsonResponse({"categories": categories})


@login_required_api
def api_dashboard(request):
    user = request.current_user
    sessions_count = ChatSession.objects.filter(user=user).count()
    saved_count = SavedPrompt.objects.filter(user=user).count()
    chats_this_week = ChatMessage.objects.filter(session__user=user, role="user").count()
    return JsonResponse(
        {
            "user_name": f"{user.first_name} {user.last_name}".strip(),
            "industry": user.industry,
            "metrics": {
                "agent_sessions": sessions_count,
                "saved_prompts": saved_count,
                "courses_enrolled": 0,
                "sessions_this_week": chats_this_week,
            },
        }
    )


# Blended token cost estimate: $0.50/1M tokens (gpt-4.1-mini rough average of input $0.40 + output $1.60).
# TODO: split ChatMessage.token_count into input/output columns for accurate cost attribution.
_BLENDED_COST_PER_TOKEN = 0.50 / 1_000_000


@csrf_exempt
@login_required_api
def api_pulse(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    now = dj_tz.now()
    today = now.date()
    day_ago = now - timedelta(hours=24)
    week_ago = now - timedelta(days=7)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    signups_today = CustomUser.objects.filter(created_at__date=today).count()
    signups_week = CustomUser.objects.filter(created_at__gte=week_ago).count()

    active_today = CustomUser.objects.filter(last_active_at__gte=day_ago).count()
    active_week = CustomUser.objects.filter(last_active_at__gte=week_ago).count()

    messages_today = ChatMessage.objects.filter(role="user", created_at__date=today).count()
    messages_week = ChatMessage.objects.filter(role="user", created_at__gte=week_ago).count()

    tokens_today = (
        ChatMessage.objects.filter(role="assistant", created_at__date=today)
        .aggregate(total=Sum("token_count"))["total"] or 0
    )
    tokens_month = (
        ChatMessage.objects.filter(role="assistant", created_at__gte=month_start)
        .aggregate(total=Sum("token_count"))["total"] or 0
    )
    cost_today = round(tokens_today * _BLENDED_COST_PER_TOKEN, 4)
    cost_month = round(tokens_month * _BLENDED_COST_PER_TOKEN, 4)

    # NOTE: errors_today uses a rolling 24-hour window (not a calendar day) so the
    # PULSE tile stays responsive regardless of wall-clock time. The OPERATIONS
    # summary endpoint (/api/admin/ops/summary) uses calendar-day counts instead,
    # matching the rest of its per-day aggregation logic.
    errors_today = ErrorLog.objects.filter(created_at__gte=day_ago).count()
    error_by_type = {
        row["error_type"]: row["n"]
        for row in ErrorLog.objects.filter(created_at__gte=day_ago)
        .values("error_type")
        .annotate(n=Count("id"))
    }

    total_users = CustomUser.objects.count()
    completed_users = CustomUser.objects.exclude(display_name="").count()
    profile_rate = round(completed_users / total_users * 100) if total_users else 0

    return JsonResponse({
        "signups": {"today": signups_today, "week": signups_week},
        "active_users": {"today": active_today, "week": active_week},
        "messages": {"today": messages_today, "week": messages_week},
        "tokens": {
            "today": tokens_today,
            "month": tokens_month,
            "estimated_cost_usd_today": cost_today,
            "estimated_cost_usd_month": cost_month,
        },
        "errors": {"today": errors_today, "by_type": error_by_type},
        "profile_completion": {
            "rate": profile_rate,
            "completed": completed_users,
            "total": total_users,
        },
    })


def _parse_active_after(value):
    """Parse active_after param: relative '24h'/'7d' or ISO datetime string."""
    now = dj_tz.now()
    if value.endswith("h"):
        try:
            return now - timedelta(hours=int(value[:-1]))
        except ValueError:
            return None
    if value.endswith("d"):
        try:
            return now - timedelta(days=int(value[:-1]))
        except ValueError:
            return None
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            from django.utils.timezone import make_aware
            dt = make_aware(dt)
        return dt
    except ValueError:
        return None


def _parse_since(value, now):
    """Parse a ?since= param into an aware datetime.

    Accepts: '24h', 'Nd' (e.g. '7d', '30d'), or an ISO date string ('YYYY-MM-DD').
    Returns None if unparseable (callers should default to a sensible fallback).
    """
    if value == "24h":
        return now - timedelta(hours=24)
    if value.endswith("d"):
        try:
            return now - timedelta(days=int(value[:-1]))
        except ValueError:
            pass
    try:
        from datetime import date as _date
        d = _date.fromisoformat(value)
        return dj_tz.make_aware(datetime(d.year, d.month, d.day))
    except (ValueError, AttributeError):
        pass
    return None


def _relative_time(dt, now):
    """Return a human-readable relative timestamp string."""
    seconds = int((now - dt).total_seconds())
    if seconds < 60:
        return "just now"
    if seconds < 3600:
        m = seconds // 60
        return f"{m} minute{'s' if m != 1 else ''} ago"
    if seconds < 86400:
        h = seconds // 3600
        return f"{h} hour{'s' if h != 1 else ''} ago"
    d = seconds // 86400
    return f"{d} day{'s' if d != 1 else ''} ago"


def _build_user_queryset(params):
    qs = CustomUser.objects.annotate(
        message_count=Count(
            "chat_sessions__messages",
            filter=Q(chat_sessions__messages__role="user"),
            distinct=True,
        ),
        session_count=Count("chat_sessions", distinct=True),
    )

    search = params.get("search", "").strip()
    if search:
        qs = qs.filter(
            Q(email__icontains=search)
            | Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(display_name__icontains=search)
            | Q(role__icontains=search)
        )

    industry = params.get("industry", "").strip().lower()
    if industry:
        qs = qs.filter(industry__iexact=INDUSTRY_MAP.get(industry, industry))

    has_profile = params.get("has_profile", "").strip().lower()
    if has_profile == "true":
        qs = qs.exclude(display_name="")
    elif has_profile == "false":
        qs = qs.filter(display_name="")

    signup_after = params.get("signup_after", "").strip()
    if signup_after:
        try:
            qs = qs.filter(created_at__date__gte=datetime.fromisoformat(signup_after).date())
        except ValueError:
            pass

    signup_before = params.get("signup_before", "").strip()
    if signup_before:
        try:
            qs = qs.filter(created_at__date__lte=datetime.fromisoformat(signup_before).date())
        except ValueError:
            pass

    active_after = params.get("active_after", "").strip()
    if active_after:
        cutoff = _parse_active_after(active_after)
        if cutoff:
            qs = qs.filter(last_active_at__gte=cutoff)

    is_suspended = params.get("is_suspended", "").strip().lower()
    if is_suspended == "true":
        qs = qs.filter(is_suspended=True)
    elif is_suspended == "false":
        qs = qs.filter(is_suspended=False)

    is_admin_filter = params.get("is_admin", "").strip().lower()
    if is_admin_filter in ("true", "false"):
        # Django auth bridge lookup — one extra query; intentionally more expensive than other filters.
        DjangoUser = get_user_model()
        su_qs = DjangoUser.objects.filter(is_superuser=True)
        admin_ids = {v.strip().lower() for v in su_qs.values_list("email", flat=True)}
        admin_ids |= {v.strip().lower() for v in su_qs.values_list("username", flat=True)}
        if is_admin_filter == "true":
            qs = qs.filter(email__in=admin_ids)
        else:
            qs = qs.exclude(email__in=admin_ids)

    sort_map = {
        "signup_desc":   ["-created_at"],
        "signup_asc":    ["created_at"],
        "active_desc":   ["-last_active_at"],
        "active_asc":    ["last_active_at"],
        "messages_desc": ["-message_count"],
        "name_asc":      ["first_name", "last_name"],
    }
    qs = qs.order_by(*sort_map.get(params.get("sort", "signup_desc").strip(), ["-created_at"]))
    return qs


def _serialize_user_item(user, superuser_emails, superuser_usernames):
    email = (user.email or "").strip().lower()
    is_admin = email in superuser_emails or email in superuser_usernames
    if not is_admin and email.endswith("@local.user"):
        is_admin = email.split("@", 1)[0] in superuser_usernames

    def iso(dt):
        return dt.isoformat().replace("+00:00", "Z") if dt else None

    return {
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "display_name": user.display_name,
        "role": user.role,
        "industry": user.industry,
        "date_joined": iso(user.created_at),
        "last_active_at": iso(user.last_active_at),
        "is_suspended": user.is_suspended,
        "is_admin": is_admin,
        "message_count": getattr(user, "message_count", 0) or 0,
        "session_count": getattr(user, "session_count", 0) or 0,
        "has_profile": bool(user.display_name),
    }


def _serialize_user_full(target, superuser_emails, superuser_usernames):
    """All-fields user serialization for admin detail/edit responses.

    Unlike _serialize_user_item (which carries annotated counts from the list
    queryset), this version includes all 18 profile fields. message_count and
    session_count are computed separately by the detail endpoint and surfaced
    under the 'stats' key, not here.
    """
    email = (target.email or "").strip().lower()
    is_admin = email in superuser_emails or email in superuser_usernames
    if not is_admin and email.endswith("@local.user"):
        is_admin = email.split("@", 1)[0] in superuser_usernames

    def iso(dt):
        return dt.isoformat().replace("+00:00", "Z") if dt else None

    return {
        "id": target.id,
        "email": target.email,
        "first_name": target.first_name,
        "last_name": target.last_name,
        "is_suspended": target.is_suspended,
        "is_admin": is_admin,
        "date_joined": iso(target.created_at),
        "last_active_at": iso(target.last_active_at),
        "industry": target.industry,
        "display_name": target.display_name,
        "role": target.role,
        "company_size": target.company_size,
        "years_experience": target.years_experience,
        "timezone": target.timezone,
        "communication_style": target.communication_style,
        "response_length": target.response_length,
        "expertise_level": target.expertise_level,
        "formality": target.formality,
        "emoji_use": target.emoji_use,
        "pushback_style": target.pushback_style,
        "explanation_style": target.explanation_style,
        "clarifying_questions": target.clarifying_questions,
        "expertise_areas": target.expertise_areas,
        "current_focus": target.current_focus,
        "things_to_avoid": target.things_to_avoid,
        "about_me": target.about_me,
    }


@csrf_exempt
@login_required_api
def api_admin_users(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    qs = _build_user_queryset(request.GET)

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (ValueError, TypeError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.GET.get("page_size", 50))))
    except (ValueError, TypeError):
        page_size = 50

    total = qs.count()
    pages = max(1, (total + page_size - 1) // page_size)
    users_page = list(qs[(page - 1) * page_size:(page - 1) * page_size + page_size])

    DjangoUser = get_user_model()
    superusers = list(DjangoUser.objects.filter(is_superuser=True).values("email", "username"))
    superuser_emails = {su["email"].strip().lower() for su in superusers}
    superuser_usernames = {su["username"].strip().lower() for su in superusers}

    return JsonResponse({
        "items": [_serialize_user_item(u, superuser_emails, superuser_usernames) for u in users_page],
        "page": page,
        "page_size": page_size,
        "total": total,
        "pages": pages,
    })


@csrf_exempt
@login_required_api
def api_admin_users_export(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    qs = _build_user_queryset(request.GET)
    total = qs.count()

    if total > 10_000:
        return JsonResponse(
            {
                "error": (
                    f"Export exceeds the 10,000-user limit ({total:,} matched). "
                    "Apply tighter filters to narrow down. "
                    "Streaming exports for larger sets are a planned future enhancement."
                )
            },
            status=413,
        )

    DjangoUser = get_user_model()
    superusers = list(DjangoUser.objects.filter(is_superuser=True).values("email", "username"))
    superuser_emails = {su["email"].strip().lower() for su in superusers}
    superuser_usernames = {su["username"].strip().lower() for su in superusers}

    today_str = dj_tz.now().strftime("%Y-%m-%d")
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="users-{today_str}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "id", "email", "first_name", "last_name", "display_name", "role",
        "industry", "company_size", "years_experience", "date_joined",
        "last_active_at", "is_suspended", "is_admin", "message_count",
        "session_count", "has_profile", "expertise_areas",
    ])
    for user in qs.iterator():
        item = _serialize_user_item(user, superuser_emails, superuser_usernames)
        expertise = ", ".join(user.expertise_areas) if isinstance(user.expertise_areas, list) else ""
        writer.writerow([
            item["id"], item["email"], item["first_name"], item["last_name"],
            item["display_name"], item["role"], item["industry"],
            user.company_size, user.years_experience,
            item["date_joined"], item["last_active_at"],
            item["is_suspended"], item["is_admin"],
            item["message_count"], item["session_count"],
            item["has_profile"], expertise,
        ])

    active_filters = {k: v for k, v in request.GET.items() if v and k not in ("page", "page_size")}
    record_admin_action(
        admin=request.current_user,
        action_type="user.export",
        target_type="user",
        metadata={"count": total, "filters": active_filters},
    )
    return response


# ── User detail (GET + PATCH) ───────────────────────────────────────────────

@csrf_exempt
@login_required_api
def api_admin_user_detail(request, user_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    target = CustomUser.objects.filter(id=user_id).first()
    if not target:
        return JsonResponse({"error": "User not found"}, status=404)

    if request.method == "GET":
        DjangoUser = get_user_model()
        superusers = list(DjangoUser.objects.filter(is_superuser=True).values("email", "username"))
        superuser_emails = {su["email"].strip().lower() for su in superusers}
        superuser_usernames = {su["username"].strip().lower() for su in superusers}

        now = dj_tz.now()
        thirty_ago = now - timedelta(days=30)
        seven_ago = now - timedelta(days=7)

        message_count = ChatMessage.objects.filter(session__user=target, role="user").count()
        session_count = ChatSession.objects.filter(user=target).count()

        top_agents = list(
            ChatSession.objects.filter(user=target)
            .values("agent_name")
            .annotate(count=Count("id"))
            .order_by("-count")[:5]
        )

        top_prompts_qs = (
            SavedPrompt.objects.filter(user=target)
            .select_related("prompt")
            .annotate(total_saves=Count("prompt__saved_by", distinct=True))
            .order_by("-created_at")[:5]
        )

        first_msg = (
            ChatMessage.objects.filter(session__user=target, role="user")
            .order_by("created_at")
            .values("created_at")
            .first()
        )

        # Distinct active days: fetch datetimes, extract unique UTC dates in Python.
        # Per-user message volume is small enough that this is efficient.
        active_dates_raw = list(
            ChatMessage.objects.filter(
                session__user=target, role="user", created_at__gte=thirty_ago
            ).values_list("created_at", flat=True)
        )
        active_days_30 = len({dt.date() for dt in active_dates_raw})

        messages_7d = ChatMessage.objects.filter(
            session__user=target, role="user", created_at__gte=seven_ago
        ).count()
        messages_30d = ChatMessage.objects.filter(
            session__user=target, role="user", created_at__gte=thirty_ago
        ).count()

        recent_sessions = list(
            ChatSession.objects.filter(user=target)
            .annotate(msg_count=Count("messages"))
            .order_by("-updated_at")[:20]
        )

        def iso(dt):
            return dt.isoformat().replace("+00:00", "Z") if dt else None

        return JsonResponse({
            "user": _serialize_user_full(target, superuser_emails, superuser_usernames),
            "stats": {
                "message_count": message_count,
                "session_count": session_count,
                "top_agents": [
                    {"name": a["agent_name"], "count": a["count"]} for a in top_agents
                ],
                "top_prompts": [
                    {"title": sp.prompt.title, "saved_count": sp.total_saves}
                    for sp in top_prompts_qs
                ],
                "engagement": {
                    "first_active": iso(first_msg["created_at"]) if first_msg else None,
                    "last_active_at": iso(target.last_active_at),
                    "active_days_last_30": active_days_30,
                    "messages_last_7d": messages_7d,
                    "messages_last_30d": messages_30d,
                },
            },
            "sessions": [
                {
                    "id": s.id,
                    "title": s.title,
                    "agent_name": s.agent_name,
                    "industry": s.industry,
                    "message_count": s.msg_count,
                    "created_at": iso(s.created_at),
                    "updated_at": iso(s.updated_at),
                }
                for s in recent_sessions
            ],
        })

    if request.method == "PATCH":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        updates, errors = _validate_and_build_profile_updates(payload)

        if "is_suspended" in payload:
            val = payload["is_suspended"]
            if not isinstance(val, bool):
                errors["is_suspended"] = "Must be a boolean."
            elif val and target.id == request.current_user.id:
                errors["is_suspended"] = "Cannot suspend yourself."
            elif val and _is_admin_user(target):
                errors["is_suspended"] = "Cannot suspend another admin."
            else:
                updates["is_suspended"] = val

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        if not updates:
            return JsonResponse({"success": True, "changed_fields": []})

        changed_fields = list(updates.keys())
        for field, value in updates.items():
            setattr(target, field, value)
        target.save(update_fields=changed_fields + ["updated_at"])

        record_admin_action(
            admin=request.current_user,
            action_type="user.edit_profile",
            target_type="user",
            target_id=target.id,
            metadata={"changed_fields": changed_fields},
        )

        DjangoUser = get_user_model()
        superusers = list(DjangoUser.objects.filter(is_superuser=True).values("email", "username"))
        superuser_emails = {su["email"].strip().lower() for su in superusers}
        superuser_usernames = {su["username"].strip().lower() for su in superusers}
        return JsonResponse({
            "success": True,
            "changed_fields": changed_fields,
            "user": _serialize_user_full(target, superuser_emails, superuser_usernames),
        })

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ── Suspend / Unsuspend ─────────────────────────────────────────────────────

@csrf_exempt
@login_required_api
def api_admin_user_suspend(request, user_id):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    target = CustomUser.objects.filter(id=user_id).first()
    if not target:
        return JsonResponse({"error": "User not found"}, status=404)
    if target.id == request.current_user.id:
        return JsonResponse({"error": "Cannot suspend yourself."}, status=400)
    if _is_admin_user(target):
        return JsonResponse({"error": "Cannot suspend another admin."}, status=400)
    if target.is_suspended:
        return JsonResponse({"error": "User is already suspended."}, status=400)

    target.is_suspended = True
    target.save(update_fields=["is_suspended", "updated_at"])

    record_admin_action(
        admin=request.current_user,
        action_type="user.suspend",
        target_type="user",
        target_id=target.id,
        metadata={},
    )
    return JsonResponse({"ok": True})


@csrf_exempt
@login_required_api
def api_admin_user_unsuspend(request, user_id):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    target = CustomUser.objects.filter(id=user_id).first()
    if not target:
        return JsonResponse({"error": "User not found"}, status=404)
    if not target.is_suspended:
        return JsonResponse({"error": "User is not suspended."}, status=400)

    target.is_suspended = False
    target.save(update_fields=["is_suspended", "updated_at"])

    record_admin_action(
        admin=request.current_user,
        action_type="user.unsuspend",
        target_type="user",
        target_id=target.id,
        metadata={},
    )
    return JsonResponse({"ok": True})


# ── Impersonation ───────────────────────────────────────────────────────────

@csrf_exempt
@login_required_api
def api_admin_user_impersonate(request, user_id):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    if "_impersonated_by" in request.session:
        return JsonResponse(
            {"error": "Already in an impersonation session. Stop the current one first."},
            status=400,
        )

    target = CustomUser.objects.filter(id=user_id).first()
    if not target:
        return JsonResponse({"error": "User not found"}, status=404)
    if target.id == request.current_user.id:
        return JsonResponse({"error": "Cannot impersonate yourself."}, status=400)
    if _is_admin_user(target):
        return JsonResponse({"error": "Cannot impersonate other admins."}, status=400)
    if target.is_suspended:
        return JsonResponse({"error": "Cannot impersonate a suspended user."}, status=400)

    request.session["_impersonated_by"] = request.current_user.id
    request.session["_impersonation_started_at"] = (
        dj_tz.now().isoformat().replace("+00:00", "Z")
    )
    login_user(request, target)  # sets custom_user_id = target.id

    record_admin_action(
        admin=request.current_user,
        action_type="user.impersonate",
        target_type="user",
        target_id=target.id,
        metadata={"target_user_id": target.id, "target_email": target.email},
    )
    return JsonResponse({"ok": True, "redirect": "/dashboard/"})


@csrf_exempt
@login_required_api
def api_admin_stop_impersonation(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if "_impersonated_by" not in request.session:
        return JsonResponse({"error": "Not currently impersonating."}, status=400)

    impersonated_user_id = request.session.get("custom_user_id")  # capture before restore
    admin_id = request.session.get("_impersonated_by")
    _do_restore_impersonation(request)  # session now points back to admin

    admin_user = CustomUser.objects.filter(id=admin_id).first()
    if admin_user:
        record_admin_action(
            admin=admin_user,
            action_type="user.stop_impersonation",
            target_type="user",
            target_id=impersonated_user_id,
            metadata={"target_user_id": impersonated_user_id},
        )
    return JsonResponse({"ok": True, "redirect": "/admin-dashboard/"})


# ---------------------------------------------------------------------------
# OPERATIONS endpoints
# ---------------------------------------------------------------------------

@csrf_exempt
@login_required_api
def api_admin_ops_summary(request):
    """OPERATIONS — summary stats card.

    "today" uses calendar-day filtering; "week" uses a rolling 7-day window.
    Both differ from /api/admin/pulse which uses a rolling 24-hour window for
    errors so its tile stays accurate regardless of wall-clock time.
    """
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    now = dj_tz.now()
    today = now.date()
    day_ago = now - timedelta(hours=24)
    week_ago = now - timedelta(days=7)

    # today (calendar day)
    messages_today = ChatMessage.objects.filter(role="user", created_at__date=today).count()
    tokens_today = (
        ChatMessage.objects.filter(role="assistant", created_at__date=today)
        .aggregate(total=Sum("token_count"))["total"] or 0
    )
    errors_today = ErrorLog.objects.filter(created_at__date=today).count()
    admin_actions_today = AdminAction.objects.filter(created_at__date=today).count()

    # week (rolling 7 days)
    messages_week = ChatMessage.objects.filter(role="user", created_at__gte=week_ago).count()
    tokens_week = (
        ChatMessage.objects.filter(role="assistant", created_at__gte=week_ago)
        .aggregate(total=Sum("token_count"))["total"] or 0
    )
    errors_week = ErrorLog.objects.filter(created_at__gte=week_ago).count()
    admin_actions_week = AdminAction.objects.filter(created_at__gte=week_ago).count()

    # health checks
    checks = []

    try:
        CustomUser.objects.count()
        checks.append({"name": "Database", "status": "ok", "detail": "Connected"})
    except Exception as exc:
        checks.append({"name": "Database", "status": "error", "detail": str(exc)[:120]})

    openai_errors = ErrorLog.objects.filter(error_type__startswith="openai_", created_at__gte=day_ago).count()
    oe_status = "ok" if openai_errors == 0 else ("warning" if openai_errors <= 5 else "error")
    checks.append({
        "name": "OpenAI errors (last 24h)",
        "status": oe_status,
        "detail": f"{openai_errors} error{'s' if openai_errors != 1 else ''}",
    })

    audit_count = AdminAction.objects.count()
    checks.append({"name": "Audit log size", "status": "ok", "detail": f"{audit_count} entr{'ies' if audit_count != 1 else 'y'}"})

    error_count = ErrorLog.objects.count()
    checks.append({"name": "Error log size", "status": "ok", "detail": f"{error_count} entr{'ies' if error_count != 1 else 'y'}"})

    status_rank = {"ok": 0, "warning": 1, "error": 2}
    overall = max(checks, key=lambda c: status_rank.get(c["status"], 0))["status"]

    return JsonResponse({
        "health": {"status": overall, "checks": checks},
        "today": {
            "messages_sent": messages_today,
            "tokens_used": tokens_today,
            "estimated_cost_usd": round(tokens_today * _BLENDED_COST_PER_TOKEN, 6),
            "errors": errors_today,
            "admin_actions": admin_actions_today,
        },
        "week": {
            "messages_sent": messages_week,
            "tokens_used": tokens_week,
            "estimated_cost_usd": round(tokens_week * _BLENDED_COST_PER_TOKEN, 6),
            "errors": errors_week,
            "admin_actions": admin_actions_week,
        },
    })


@csrf_exempt
@login_required_api
def api_admin_ops_token_usage(request):
    """OPERATIONS — daily token usage for chart rendering.

    Returns exactly ?days=N entries (default 30), newest last, with zero backfill
    so the chart never has gaps.  The loop runs from (today - N + 1) to today
    inclusive, giving precisely N days with no off-by-one.
    """
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    now = dj_tz.now()
    today = now.date()
    days_back = min(365, max(1, int(request.GET.get("days", 30))))
    # start_date is inclusive; today - (N-1) days gives exactly N days in range
    start_date = today - timedelta(days=days_back - 1)

    rows = (
        ChatMessage.objects.filter(role="assistant", created_at__date__gte=start_date)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(tokens=Sum("token_count"), messages=Count("id"))
        .order_by("day")
    )

    by_day = {row["day"]: {"tokens": row["tokens"] or 0, "messages": row["messages"]} for row in rows}

    result = []
    current = start_date
    while current <= today:
        entry = by_day.get(current, {"tokens": 0, "messages": 0})
        result.append({
            "date": current.isoformat(),
            "tokens": entry["tokens"],
            "messages": entry["messages"],
            "cost_usd": round(entry["tokens"] * _BLENDED_COST_PER_TOKEN, 6),
        })
        current += timedelta(days=1)

    total_tokens = sum(d["tokens"] for d in result)
    total_messages = sum(d["messages"] for d in result)

    return JsonResponse({
        "days": result,
        "totals": {
            "tokens": total_tokens,
            "messages": total_messages,
            "cost_usd": round(total_tokens * _BLENDED_COST_PER_TOKEN, 6),
        },
    })


@csrf_exempt
@login_required_api
def api_admin_ops_audit(request):
    """OPERATIONS — paginated admin audit log."""
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    now = dj_tz.now()
    since_raw = request.GET.get("since", "30d")
    since_dt = _parse_since(since_raw, now) or (now - timedelta(days=30))
    action_type = request.GET.get("action_type", "").strip()
    q = request.GET.get("q", "").strip()

    page = max(1, int(request.GET.get("page", 1)))
    page_size = min(100, max(1, int(request.GET.get("page_size", 50))))

    qs = (
        AdminAction.objects.filter(created_at__gte=since_dt)
        .select_related("admin")
        .order_by("-created_at")
    )
    if action_type:
        qs = qs.filter(action_type__icontains=action_type)
    if q:
        qs = qs.filter(Q(action_type__icontains=q) | Q(target_type__icontains=q))

    total = qs.count()
    pages = max(1, (total + page_size - 1) // page_size)
    items_qs = qs[(page - 1) * page_size: page * page_size]

    items = [
        {
            "id": a.id,
            "admin_id": a.admin.id if a.admin else None,
            "admin_email": a.admin.email if a.admin else None,
            "admin_display_name": a.admin.display_name if a.admin else None,
            "action_type": a.action_type,
            "target_type": a.target_type,
            "target_id": a.target_id,
            "metadata": a.metadata,
            "created_at": a.created_at.isoformat(),
            "created_at_relative": _relative_time(a.created_at, now),
        }
        for a in items_qs
    ]

    return JsonResponse({"items": items, "page": page, "page_size": page_size, "total": total, "pages": pages})


@csrf_exempt
@login_required_api
def api_admin_ops_errors(request):
    """OPERATIONS — paginated error log."""
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    now = dj_tz.now()
    since_raw = request.GET.get("since", "7d")
    since_dt = _parse_since(since_raw, now) or (now - timedelta(days=7))
    error_type = request.GET.get("error_type", "").strip()
    q = request.GET.get("q", "").strip()

    page = max(1, int(request.GET.get("page", 1)))
    page_size = min(100, max(1, int(request.GET.get("page_size", 50))))

    qs = (
        ErrorLog.objects.filter(created_at__gte=since_dt)
        .select_related("user")
        .order_by("-created_at")
    )
    if error_type:
        qs = qs.filter(error_type__icontains=error_type)
    if q:
        qs = qs.filter(Q(error_type__icontains=q) | Q(message__icontains=q))

    total = qs.count()
    pages = max(1, (total + page_size - 1) // page_size)
    items_qs = qs[(page - 1) * page_size: page * page_size]

    items = [
        {
            "id": e.id,
            "error_type": e.error_type,
            "message": e.message,
            "user_email": e.user.email if e.user else None,
            "user_display_name": e.user.display_name if e.user else None,
            "metadata": e.metadata,
            "created_at": e.created_at.isoformat(),
            "created_at_relative": _relative_time(e.created_at, now),
        }
        for e in items_qs
    ]

    return JsonResponse({"items": items, "page": page, "page_size": page_size, "total": total, "pages": pages})


@login_required_api
def api_prompts(request):
    ensure_seed_prompts()
    base_queryset = Prompt.objects.filter(status="approved")
    mine_only = request.GET.get("mine", "").strip() == "1"
    saved_only = request.GET.get("saved", "").strip() == "1"
    q = request.GET.get("q", "").strip()
    industry = request.GET.get("industry", "").strip().lower()
    category = request.GET.get("category", "").strip()
    queryset = Prompt.objects.filter(created_by=request.current_user) if mine_only else base_queryset
    if q:
        queryset = queryset.filter(Q(title__icontains=q) | Q(body__icontains=q))
    if industry and industry != "all":
        queryset = queryset.filter(industry__iexact=industry)
    if category and category != "All Categories":
        queryset = queryset.filter(category__iexact=category)

    user_saved_ids = set(
        SavedPrompt.objects.filter(user=request.current_user).values_list("prompt_id", flat=True)
    )
    if saved_only:
        queryset = queryset.filter(id__in=user_saved_ids)
    payload = [
        {
            "id": prompt.id,
            "title": prompt.title,
            "body": prompt.body,
            "industry": prompt.industry,
            "category": prompt.category,
            "uses": prompt.usage_count,
            "saved": prompt.id in user_saved_ids,
            "status": prompt.status,
        }
        for prompt in queryset.order_by("-usage_count", "-created_at")
    ]
    stats = {
        "total_prompts": base_queryset.count(),
        "total_industries": base_queryset.values("industry").distinct().count(),
        "total_categories": base_queryset.values("category").distinct().count(),
    }
    library = {
        "saved_count": len(user_saved_ids),
        "submissions_count": Prompt.objects.filter(created_by=request.current_user).count(),
    }
    return JsonResponse({"items": payload, "stats": stats, "library": library})


@csrf_exempt
@login_required_api
def api_submit_prompt(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)
    title = str(payload.get("title", "")).strip()
    body = str(payload.get("body", "")).strip()
    industry = str(payload.get("industry", "")).strip().lower()
    category = str(payload.get("category", "")).strip()
    if not title or not body or not industry:
        return JsonResponse({"error": "title, body, industry are required"}, status=400)

    prompt = Prompt.objects.create(
        title=title,
        body=body,
        industry=industry,
        category=category or "General",
        status="pending",
        created_by=request.current_user,
    )
    return JsonResponse({"id": prompt.id, "status": prompt.status}, status=201)


@csrf_exempt
@login_required_api
def api_toggle_save_prompt(request, prompt_id):
    prompt = Prompt.objects.filter(id=prompt_id).first()
    if not prompt:
        return JsonResponse({"error": "Prompt not found"}, status=404)
    if request.method == "POST":
        SavedPrompt.objects.get_or_create(user=request.current_user, prompt=prompt)
        return JsonResponse({"saved": True})
    if request.method == "DELETE":
        SavedPrompt.objects.filter(user=request.current_user, prompt=prompt).delete()
        return JsonResponse({"saved": False})
    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required_api
def api_chat_sessions(request):
    sessions = ChatSession.objects.filter(user=request.current_user).order_by("-updated_at")
    return JsonResponse(
        {
            "items": [
                {
                    "id": session.id,
                    "title": session.title,
                    "agent_name": session.agent_name,
                    "industry": session.industry,
                    "updated_at": session.updated_at.isoformat(),
                }
                for session in sessions
            ]
        }
    )


@csrf_exempt
@login_required_api
def api_create_chat_session(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    payload = get_request_json(request)
    if payload is None:
        payload = {}
    session = ChatSession.objects.create(
        user=request.current_user,
        title=payload.get("title") or "New chat",
        agent_name=payload.get("agent_name") or "Contract Review Agent",
        industry=payload.get("industry") or request.current_user.industry or "legal",
    )
    return JsonResponse({"id": session.id, "title": session.title}, status=201)


@login_required_api
def api_chat_messages(request, session_id):
    session = ChatSession.objects.filter(id=session_id, user=request.current_user).first()
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)
    items = [
        {
            "id": message.id,
            "role": message.role,
            "content": message.content,
            "token_count": message.token_count,
            "created_at": message.created_at.isoformat(),
        }
        for message in session.messages.order_by("created_at")
    ]
    return JsonResponse({"session_id": session.id, "items": items})


@csrf_exempt
@login_required_api
def api_chat_upload_file(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    file_name = uploaded_file.name or "uploaded_file"
    file_bytes = uploaded_file.read()
    if not file_bytes:
        return JsonResponse({"error": "Uploaded file is empty"}, status=400)

    extracted_text = _extract_text_from_uploaded_file(file_name, file_bytes)
    cloud_url = _upload_to_cloudinary(file_name, file_bytes)
    summary_lines = [
        f"File: {file_name}",
        f"Cloud URL: {cloud_url or 'Not available'}",
    ]
    if extracted_text:
        summary_lines.append("")
        summary_lines.append("Extracted content:")
        summary_lines.append(extracted_text[:12000])
    else:
        summary_lines.append("")
        summary_lines.append(
            "No text could be extracted automatically. Please provide instructions for this file."
        )
    return JsonResponse(
        {
            "file_name": file_name,
            "cloud_url": cloud_url,
            "extracted_text": extracted_text,
            "message_text": "\n".join(summary_lines),
        }
    )


@csrf_exempt
@login_required_api
def api_send_chat_message(request, session_id):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    session = ChatSession.objects.filter(id=session_id, user=request.current_user).first()
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)
    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)

    content = str(payload.get("content", "")).strip()
    if not content:
        return JsonResponse({"error": "content is required"}, status=400)

    request.current_user.last_active_at = dj_tz.now()
    request.current_user.save(update_fields=["last_active_at"])

    user_message = ChatMessage.objects.create(session=session, role="user", content=content)
    history = list(session.messages.order_by("created_at").values("role", "content"))
    agent = Agent.objects.filter(name=session.agent_name).first()
    agent_prompts = list(agent.prompts.all()) if agent else []
    system_content = build_full_system_prompt(agent, agent_prompts, request.current_user)
    trimmed = trim_history_to_budget(history, system_content)

    if estimate_messages_tokens(system_content, trimmed) > MAX_REQUEST_TOKENS:
        user_message.delete()
        return JsonResponse(
            {
                "error": "This message is too long. Try splitting it into smaller messages.",
                "retry_content": content,
            },
            status=413,
        )

    openai_messages = [{"role": "system", "content": system_content}]
    openai_messages.extend({"role": m["role"], "content": m["content"]} for m in trimmed)

    try:
        assistant_text, total_tokens = get_openai_reply(openai_messages)
    except Exception as exc:
        user_message.delete()
        logger.exception("Chat send failed (session=%s): %s", session.id, exc)
        log_error(
            error_type="openai_chat_send",
            message=str(exc),
            user=request.current_user,
            metadata={"session_id": session.id},
        )
        return JsonResponse(
            {
                "error": "Could not generate a reply. Please try again.",
                "retry_content": content,
            },
            status=502,
        )

    assistant = ChatMessage.objects.create(
        session=session,
        role="assistant",
        content=assistant_text,
        token_count=total_tokens,
    )
    session.title = session.title if session.title != "New chat" else content[:60]
    session.save(update_fields=["title", "updated_at"])

    return JsonResponse(
        {
            "id": assistant.id,
            "role": "assistant",
            "content": assistant.content,
            "token_count": assistant.token_count,
        }
    )


@csrf_exempt
@login_required_api
def api_message_feedback(request, message_id):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    message = ChatMessage.objects.filter(id=message_id, session__user=request.current_user).first()
    if not message:
        return JsonResponse({"error": "Message not found"}, status=404)
    payload = get_request_json(request)
    if payload is None:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)
    feedback = str(payload.get("feedback", "")).strip().lower()
    if feedback not in {"up", "down"}:
        return JsonResponse({"error": "feedback must be 'up' or 'down'"}, status=400)
    message.feedback = feedback
    message.save(update_fields=["feedback"])
    return JsonResponse({"success": True, "feedback": feedback})


# ---------------------------------------------------------------------------
# INDUSTRIES — public + admin CRUD
# ---------------------------------------------------------------------------

def _serialize_industry(industry):
    return {
        "id": industry.id,
        "slug": industry.slug,
        "name": industry.name,
        "icon_class": industry.icon_class,
        "sort_order": industry.sort_order,
        "is_active": industry.is_active,
        "created_at": industry.created_at.isoformat(),
        "updated_at": industry.updated_at.isoformat(),
    }


def api_industries(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    ensure_seed_industries()
    industries = list(Industry.objects.filter(is_active=True).order_by("sort_order", "name"))
    return JsonResponse({"industries": [_serialize_industry(i) for i in industries]})


@csrf_exempt
@login_required_api
def api_admin_industries(request):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    if request.method == "GET":
        ensure_seed_industries()
        qs = Industry.objects.all()
        search = request.GET.get("search", "").strip()
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(slug__icontains=search))
        is_active_param = request.GET.get("is_active", "").strip().lower()
        if is_active_param == "true":
            qs = qs.filter(is_active=True)
        elif is_active_param == "false":
            qs = qs.filter(is_active=False)
        qs = qs.order_by("sort_order", "name")
        return JsonResponse({"items": [_serialize_industry(i) for i in qs]})

    if request.method == "POST":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        name = str(payload.get("name", "")).strip()
        errors = {}
        if not name:
            errors["name"] = "Required."
        elif len(name) > 80:
            errors["name"] = "Too long (max 80)."

        icon_class = str(payload.get("icon_class", "fa-briefcase")).strip() or "fa-briefcase"
        if len(icon_class) > 80:
            errors["icon_class"] = "Too long (max 80)."

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        slug = slugify(name)
        if not slug:
            return JsonResponse({"error": "Invalid name — cannot generate slug"}, status=400)
        if Industry.objects.filter(slug=slug).exists():
            return JsonResponse({"error": f"Industry with slug '{slug}' already exists"}, status=409)

        sort_order = _to_non_negative_int(payload.get("sort_order", 100), default=100)
        is_active = payload.get("is_active", True)
        if not isinstance(is_active, bool):
            is_active = True

        industry = Industry.objects.create(
            name=name,
            slug=slug,
            icon_class=icon_class,
            sort_order=sort_order,
            is_active=is_active,
        )
        record_admin_action(
            admin=request.current_user,
            action_type="industry.create",
            target_type="industry",
            target_id=industry.id,
            metadata={"name": industry.name, "slug": industry.slug},
        )
        return JsonResponse(_serialize_industry(industry), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required_api
def api_admin_industry_detail(request, industry_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    industry = Industry.objects.filter(id=industry_id).first()
    if not industry:
        return JsonResponse({"error": "Industry not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(_serialize_industry(industry))

    if request.method == "DELETE":
        agent_count = Agent.objects.filter(industry__iexact=industry.slug).count()
        prompt_count = Prompt.objects.filter(industry__iexact=industry.slug).count()
        if agent_count or prompt_count:
            return JsonResponse(
                {
                    "error": (
                        f"Cannot delete '{industry.name}': {agent_count} agent(s) and "
                        f"{prompt_count} prompt(s) use this industry. "
                        "Reassign them or set is_active=false instead."
                    )
                },
                status=409,
            )
        industry_name = industry.name
        industry_slug = industry.slug
        industry.delete()
        record_admin_action(
            admin=request.current_user,
            action_type="industry.delete",
            target_type="industry",
            target_id=industry_id,
            metadata={"name": industry_name, "slug": industry_slug},
        )
        return JsonResponse({"deleted": True})

    if request.method == "PATCH":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        updates = {}

        if "name" in payload:
            name = str(payload["name"]).strip()
            if not name:
                errors["name"] = "Cannot be empty."
            elif len(name) > 80:
                errors["name"] = "Too long (max 80)."
            else:
                updates["name"] = name
                new_slug = slugify(name)
                if not new_slug:
                    errors["name"] = "Invalid name."
                elif Industry.objects.filter(slug=new_slug).exclude(id=industry.id).exists():
                    errors["name"] = "Another industry with this name already exists."
                else:
                    updates["slug"] = new_slug

        if "icon_class" in payload:
            val = str(payload["icon_class"]).strip()
            if len(val) > 80:
                errors["icon_class"] = "Too long (max 80)."
            else:
                updates["icon_class"] = val or "fa-briefcase"

        if "sort_order" in payload:
            updates["sort_order"] = _to_non_negative_int(payload["sort_order"], default=industry.sort_order)

        if "is_active" in payload:
            val = payload["is_active"]
            if not isinstance(val, bool):
                errors["is_active"] = "Must be a boolean."
            else:
                updates["is_active"] = val

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        changed_fields = list(updates.keys())
        if updates:
            for field, value in updates.items():
                setattr(industry, field, value)
            industry.save(update_fields=changed_fields + ["updated_at"])

        record_admin_action(
            admin=request.current_user,
            action_type="industry.update",
            target_type="industry",
            target_id=industry.id,
            metadata={"changed_fields": changed_fields},
        )
        industry.refresh_from_db()
        return JsonResponse(_serialize_industry(industry))

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ---------------------------------------------------------------------------
# EVENTS — public + admin CRUD
# ---------------------------------------------------------------------------

def _serialize_event(event):
    industry = None
    if event.industry_id:
        ind = event.industry
        industry = {"slug": ind.slug, "name": ind.name, "icon_class": ind.icon_class}
    return {
        "id": event.id,
        "title": event.title,
        "slug": event.slug,
        "description": event.description,
        "event_date": event.event_date.isoformat(),
        "event_type": event.event_type,
        "location": event.location,
        "image_url": event.image_url,
        "registration_url": event.registration_url,
        "is_featured": event.is_featured,
        "is_active": event.is_active,
        "sort_order": event.sort_order,
        "industry": industry,
        "created_at": event.created_at.isoformat(),
        "updated_at": event.updated_at.isoformat(),
    }


def api_events(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    ensure_seed_events()
    now = dj_tz.now()
    events = list(
        Event.objects.select_related("industry")
        .filter(is_active=True, event_date__gte=now)
        .order_by("sort_order", "event_date")
    )
    return JsonResponse({"events": [_serialize_event(e) for e in events]})


@csrf_exempt
@login_required_api
def api_admin_events(request):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    if request.method == "GET":
        ensure_seed_events()
        qs = Event.objects.select_related("industry").all()

        search = request.GET.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(title__icontains=search)
                | Q(description__icontains=search)
                | Q(location__icontains=search)
            )

        industry_param = request.GET.get("industry", "").strip().lower()
        if industry_param:
            qs = qs.filter(industry__slug=industry_param)

        is_active_param = request.GET.get("is_active", "").strip().lower()
        if is_active_param == "true":
            qs = qs.filter(is_active=True)
        elif is_active_param == "false":
            qs = qs.filter(is_active=False)

        sort_param = request.GET.get("sort", "date_asc")
        if sort_param == "date_desc":
            qs = qs.order_by("-event_date")
        elif sort_param == "title_asc":
            qs = qs.order_by("title")
        elif sort_param == "featured_first":
            qs = qs.order_by("-is_featured", "event_date")
        else:
            qs = qs.order_by("sort_order", "event_date")

        page = max(1, _to_non_negative_int(request.GET.get("page", 1), default=1))
        page_size = min(200, max(1, _to_non_negative_int(request.GET.get("page_size", 50), default=50)))
        total = qs.count()
        items = list(qs[(page - 1) * page_size: page * page_size])
        return JsonResponse({
            "items": [_serialize_event(e) for e in items],
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": max(1, -(-total // page_size)),
        })

    if request.method == "POST":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        title = str(payload.get("title", "")).strip()
        if not title:
            errors["title"] = "Required."
        elif len(title) > 200:
            errors["title"] = "Too long (max 200)."

        description = str(payload.get("description", "")).strip()
        if not description:
            errors["description"] = "Required."

        event_date_raw = payload.get("event_date", "")
        event_date = _parse_event_date(event_date_raw)
        if event_date is None:
            errors["event_date"] = "Required. Use ISO-8601 (e.g. 2026-05-15T14:00)."

        _url_validator = URLValidator()
        for url_field in ("image_url", "registration_url"):
            val = str(payload.get(url_field, "") or "").strip()
            if val:
                try:
                    _url_validator(val)
                except ValidationError:
                    errors[url_field] = "Enter a valid URL."

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        slug = slugify(title)
        if not slug:
            return JsonResponse({"error": "Invalid title — cannot generate slug"}, status=400)
        if Event.objects.filter(slug=slug).exists():
            return JsonResponse({"error": f"Event with slug '{slug}' already exists"}, status=409)

        event_type = str(payload.get("event_type", "") or "").strip()[:50]
        location = str(payload.get("location", "") or "").strip()[:200]
        image_url = str(payload.get("image_url", "") or "").strip()
        registration_url = str(payload.get("registration_url", "") or "").strip()
        is_featured = bool(payload.get("is_featured", False))
        is_active = payload.get("is_active", True)
        if not isinstance(is_active, bool):
            is_active = True
        sort_order = _to_non_negative_int(payload.get("sort_order", 100), default=100)

        industry = None
        industry_slug = str(payload.get("industry", "") or "").strip()
        if industry_slug:
            industry = Industry.objects.filter(slug=industry_slug).first()

        event = Event.objects.create(
            title=title,
            slug=slug,
            description=description,
            event_date=event_date,
            event_type=event_type,
            location=location,
            image_url=image_url,
            registration_url=registration_url,
            industry=industry,
            is_featured=is_featured,
            is_active=is_active,
            sort_order=sort_order,
        )
        record_admin_action(
            admin=request.current_user,
            action_type="event.create",
            target_type="event",
            target_id=event.id,
            metadata={"title": event.title, "event_type": event.event_type},
        )
        event_out = Event.objects.select_related("industry").get(id=event.id)
        return JsonResponse(_serialize_event(event_out), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required_api
def api_admin_event_detail(request, event_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    event = Event.objects.select_related("industry").filter(id=event_id).first()
    if not event:
        return JsonResponse({"error": "Event not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(_serialize_event(event))

    if request.method == "DELETE":
        title = event.title
        event.delete()
        record_admin_action(
            admin=request.current_user,
            action_type="event.delete",
            target_type="event",
            target_id=event_id,
            metadata={"title": title},
        )
        return JsonResponse({"deleted": True})

    if request.method == "PATCH":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        updates = {}

        if "title" in payload:
            title = str(payload["title"]).strip()
            if not title:
                errors["title"] = "Cannot be empty."
            elif len(title) > 200:
                errors["title"] = "Too long (max 200)."
            else:
                new_slug = slugify(title)
                if not new_slug:
                    errors["title"] = "Invalid title."
                elif Event.objects.filter(slug=new_slug).exclude(id=event.id).exists():
                    errors["title"] = "Another event with this title already exists."
                else:
                    updates["title"] = title
                    updates["slug"] = new_slug

        if "description" in payload:
            val = str(payload["description"]).strip()
            if not val:
                errors["description"] = "Cannot be empty."
            else:
                updates["description"] = val

        if "event_date" in payload:
            parsed = _parse_event_date(payload["event_date"])
            if parsed is None:
                errors["event_date"] = "Invalid date. Use ISO-8601."
            else:
                updates["event_date"] = parsed

        _url_validator = URLValidator()
        for url_field in ("image_url", "registration_url"):
            if url_field in payload:
                val = str(payload[url_field] or "").strip()
                if val:
                    try:
                        _url_validator(val)
                    except ValidationError:
                        errors[url_field] = "Enter a valid URL."
                        continue
                updates[url_field] = val

        for str_field, max_len in (("event_type", 50), ("location", 200)):
            if str_field in payload:
                updates[str_field] = str(payload[str_field] or "").strip()[:max_len]

        for bool_field in ("is_featured", "is_active"):
            if bool_field in payload:
                val = payload[bool_field]
                if not isinstance(val, bool):
                    errors[bool_field] = "Must be a boolean."
                else:
                    updates[bool_field] = val

        if "sort_order" in payload:
            updates["sort_order"] = _to_non_negative_int(payload["sort_order"], default=event.sort_order)

        if "industry" in payload:
            industry_slug = str(payload["industry"] or "").strip()
            if industry_slug:
                ind = Industry.objects.filter(slug=industry_slug).first()
                if not ind:
                    errors["industry"] = f"Industry '{industry_slug}' not found."
                else:
                    updates["industry_id"] = ind.id
            else:
                updates["industry_id"] = None

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        changed_fields = list(updates.keys())
        if updates:
            for field, value in updates.items():
                setattr(event, field, value)
            event.save(update_fields=changed_fields + ["updated_at"])

        record_admin_action(
            admin=request.current_user,
            action_type="event.update",
            target_type="event",
            target_id=event.id,
            metadata={"changed_fields": changed_fields},
        )
        event = Event.objects.select_related("industry").get(id=event.id)
        return JsonResponse(_serialize_event(event))

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ---------------------------------------------------------------------------
# TOOLS — public + admin CRUD
# ---------------------------------------------------------------------------

def _serialize_tool(tool):
    return {
        "id": tool.id,
        "name": tool.name,
        "slug": tool.slug,
        "description": tool.description,
        "icon_class": tool.icon_class,
        "external_url": tool.external_url,
        "category": tool.category,
        "is_featured": tool.is_featured,
        "is_active": tool.is_active,
        "sort_order": tool.sort_order,
        "created_at": tool.created_at.isoformat(),
        "updated_at": tool.updated_at.isoformat(),
    }


def api_tools(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    ensure_seed_tools()
    tools = list(Tool.objects.filter(is_active=True).order_by("sort_order", "name"))
    return JsonResponse({"tools": [_serialize_tool(t) for t in tools]})


@csrf_exempt
@login_required_api
def api_admin_tools(request):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    if request.method == "GET":
        ensure_seed_tools()
        qs = Tool.objects.all()

        search = request.GET.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(description__icontains=search)
                | Q(category__icontains=search)
            )

        category_param = request.GET.get("category", "").strip()
        if category_param:
            qs = qs.filter(category__iexact=category_param)

        is_active_param = request.GET.get("is_active", "").strip().lower()
        if is_active_param == "true":
            qs = qs.filter(is_active=True)
        elif is_active_param == "false":
            qs = qs.filter(is_active=False)

        sort_param = request.GET.get("sort", "name_asc")
        if sort_param == "sort_order_asc":
            qs = qs.order_by("sort_order", "name")
        elif sort_param == "featured_first":
            qs = qs.order_by("-is_featured", "sort_order", "name")
        else:
            qs = qs.order_by("name")

        page = max(1, _to_non_negative_int(request.GET.get("page", 1), default=1))
        page_size = min(200, max(1, _to_non_negative_int(request.GET.get("page_size", 50), default=50)))
        total = qs.count()
        items = list(qs[(page - 1) * page_size: page * page_size])
        return JsonResponse({
            "items": [_serialize_tool(t) for t in items],
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": max(1, -(-total // page_size)),
        })

    if request.method == "POST":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        name = str(payload.get("name", "")).strip()
        if not name:
            errors["name"] = "Required."
        elif len(name) > 120:
            errors["name"] = "Too long (max 120)."

        description = str(payload.get("description", "")).strip()
        if not description:
            errors["description"] = "Required."

        external_url = str(payload.get("external_url", "") or "").strip()
        if not external_url:
            errors["external_url"] = "Required."
        else:
            try:
                URLValidator()(external_url)
            except ValidationError:
                errors["external_url"] = "Enter a valid URL."

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        slug = slugify(name)
        if not slug:
            return JsonResponse({"error": "Invalid name — cannot generate slug"}, status=400)
        if Tool.objects.filter(slug=slug).exists():
            return JsonResponse({"error": f"Tool with slug '{slug}' already exists"}, status=409)

        icon_class = str(payload.get("icon_class", "fa-toolbox") or "fa-toolbox").strip()[:80] or "fa-toolbox"
        category = str(payload.get("category", "") or "").strip()[:80]
        is_featured = bool(payload.get("is_featured", False))
        is_active = payload.get("is_active", True)
        if not isinstance(is_active, bool):
            is_active = True
        sort_order = _to_non_negative_int(payload.get("sort_order", 100), default=100)

        tool = Tool.objects.create(
            name=name,
            slug=slug,
            description=description,
            icon_class=icon_class,
            external_url=external_url,
            category=category,
            is_featured=is_featured,
            is_active=is_active,
            sort_order=sort_order,
        )
        record_admin_action(
            admin=request.current_user,
            action_type="tool.create",
            target_type="tool",
            target_id=tool.id,
            metadata={"name": tool.name, "slug": tool.slug},
        )
        return JsonResponse(_serialize_tool(tool), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required_api
def api_admin_tool_detail(request, tool_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    tool = Tool.objects.filter(id=tool_id).first()
    if not tool:
        return JsonResponse({"error": "Tool not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(_serialize_tool(tool))

    if request.method == "DELETE":
        name = tool.name
        tool.delete()
        record_admin_action(
            admin=request.current_user,
            action_type="tool.delete",
            target_type="tool",
            target_id=tool_id,
            metadata={"name": name},
        )
        return JsonResponse({"deleted": True})

    if request.method == "PATCH":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        updates = {}

        if "name" in payload:
            name = str(payload["name"]).strip()
            if not name:
                errors["name"] = "Cannot be empty."
            elif len(name) > 120:
                errors["name"] = "Too long (max 120)."
            else:
                new_slug = slugify(name)
                if not new_slug:
                    errors["name"] = "Invalid name."
                elif Tool.objects.filter(slug=new_slug).exclude(id=tool.id).exists():
                    errors["name"] = "Another tool with this name already exists."
                else:
                    updates["name"] = name
                    updates["slug"] = new_slug

        if "description" in payload:
            val = str(payload["description"]).strip()
            if not val:
                errors["description"] = "Cannot be empty."
            else:
                updates["description"] = val

        if "external_url" in payload:
            val = str(payload["external_url"] or "").strip()
            if not val:
                errors["external_url"] = "Required."
            else:
                try:
                    URLValidator()(val)
                    updates["external_url"] = val
                except ValidationError:
                    errors["external_url"] = "Enter a valid URL."

        if "icon_class" in payload:
            val = str(payload["icon_class"] or "fa-toolbox").strip()
            updates["icon_class"] = (val or "fa-toolbox")[:80]

        if "category" in payload:
            updates["category"] = str(payload["category"] or "").strip()[:80]

        for bool_field in ("is_featured", "is_active"):
            if bool_field in payload:
                val = payload[bool_field]
                if not isinstance(val, bool):
                    errors[bool_field] = "Must be a boolean."
                else:
                    updates[bool_field] = val

        if "sort_order" in payload:
            updates["sort_order"] = _to_non_negative_int(payload["sort_order"], default=tool.sort_order)

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        changed_fields = list(updates.keys())
        if updates:
            for field, value in updates.items():
                setattr(tool, field, value)
            tool.save(update_fields=changed_fields + ["updated_at"])

        record_admin_action(
            admin=request.current_user,
            action_type="tool.update",
            target_type="tool",
            target_id=tool.id,
            metadata={"changed_fields": changed_fields},
        )
        tool.refresh_from_db()
        return JsonResponse(_serialize_tool(tool))

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ── BANNERS ───────────────────────────────────────────────────────────────────

def _serialize_banner(banner):
    return {
        "id": banner.id,
        "message": banner.message,
        "banner_type": banner.banner_type,
        "start_at": banner.start_at.isoformat(),
        "end_at": banner.end_at.isoformat(),
        "is_active": banner.is_active,
        "is_dismissible": banner.is_dismissible,
        "created_by_email": banner.created_by.email if banner.created_by else None,
        "created_at": banner.created_at.isoformat(),
        "updated_at": banner.updated_at.isoformat(),
    }


@csrf_exempt
@login_required_api
def api_banners(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    now = dj_tz.now()
    banners = list(
        Banner.objects.filter(
            is_active=True,
            start_at__lte=now,
            end_at__gte=now,
        ).order_by("-start_at")
    )
    return JsonResponse({"banners": [_serialize_banner(b) for b in banners]})


_BANNER_TYPES = {"info", "warn", "critical"}


@csrf_exempt
@login_required_api
def api_admin_banners(request):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)

    if request.method == "GET":
        qs = Banner.objects.all()

        is_active_param = request.GET.get("is_active", "").strip().lower()
        if is_active_param == "true":
            qs = qs.filter(is_active=True)
        elif is_active_param == "false":
            qs = qs.filter(is_active=False)

        banner_type_param = request.GET.get("banner_type", "").strip().lower()
        if banner_type_param in _BANNER_TYPES:
            qs = qs.filter(banner_type=banner_type_param)

        sort_param = request.GET.get("sort", "-start_at")
        if sort_param not in {"-start_at", "start_at", "-end_at", "end_at"}:
            sort_param = "-start_at"
        qs = qs.order_by(sort_param)

        return JsonResponse({"banners": [_serialize_banner(b) for b in qs]})

    if request.method == "POST":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}

        message = str(payload.get("message", "")).strip()
        if not message:
            errors["message"] = "Required."

        banner_type = str(payload.get("banner_type", "info")).strip().lower()
        if banner_type not in _BANNER_TYPES:
            errors["banner_type"] = "Must be info, warn, or critical."

        start_at = _parse_event_date(payload.get("start_at"))
        if start_at is None:
            errors["start_at"] = "Required. Use ISO-8601 (e.g. 2026-05-15T14:00)."

        end_at = _parse_event_date(payload.get("end_at"))
        if end_at is None:
            errors["end_at"] = "Required. Use ISO-8601 (e.g. 2026-05-15T14:00)."

        if start_at and end_at and start_at >= end_at:
            errors["end_at"] = "Must be after start time."

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        is_active = payload.get("is_active", True)
        if not isinstance(is_active, bool):
            is_active = True

        is_dismissible = payload.get("is_dismissible", True)
        if not isinstance(is_dismissible, bool):
            is_dismissible = True

        banner = Banner.objects.create(
            message=message,
            banner_type=banner_type,
            start_at=start_at,
            end_at=end_at,
            is_active=is_active,
            is_dismissible=is_dismissible,
            created_by=request.current_user,
        )
        record_admin_action(
            admin=request.current_user,
            action_type="banner.create",
            target_type="banner",
            target_id=banner.id,
            metadata={"banner_type": banner.banner_type},
        )
        return JsonResponse(_serialize_banner(banner), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required_api
def api_admin_banner_detail(request, banner_id):
    if not _is_admin_user(request.current_user):
        return JsonResponse({"error": "Admin access required"}, status=403)
    banner = Banner.objects.filter(id=banner_id).first()
    if not banner:
        return JsonResponse({"error": "Banner not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(_serialize_banner(banner))

    if request.method == "DELETE":
        banner_type = banner.banner_type
        banner.delete()
        record_admin_action(
            admin=request.current_user,
            action_type="banner.delete",
            target_type="banner",
            target_id=banner_id,
            metadata={"banner_type": banner_type},
        )
        return JsonResponse({"deleted": True})

    if request.method == "PATCH":
        payload = get_request_json(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload"}, status=400)

        errors = {}
        updates = {}

        if "message" in payload:
            val = str(payload["message"]).strip()
            if not val:
                errors["message"] = "Cannot be empty."
            else:
                updates["message"] = val

        if "banner_type" in payload:
            val = str(payload["banner_type"]).strip().lower()
            if val not in _BANNER_TYPES:
                errors["banner_type"] = "Must be info, warn, or critical."
            else:
                updates["banner_type"] = val

        if "start_at" in payload:
            dt = _parse_event_date(payload["start_at"])
            if dt is None:
                errors["start_at"] = "Must be a valid ISO-8601 datetime."
            else:
                updates["start_at"] = dt

        if "end_at" in payload:
            dt = _parse_event_date(payload["end_at"])
            if dt is None:
                errors["end_at"] = "Must be a valid ISO-8601 datetime."
            else:
                updates["end_at"] = dt

        if "start_at" in updates or "end_at" in updates:
            new_start = updates.get("start_at", banner.start_at)
            new_end = updates.get("end_at", banner.end_at)
            if new_start >= new_end:
                errors["end_at"] = "Must be after start time."

        for bool_field in ("is_active", "is_dismissible"):
            if bool_field in payload:
                val = payload[bool_field]
                if not isinstance(val, bool):
                    errors[bool_field] = "Must be a boolean."
                else:
                    updates[bool_field] = val

        if errors:
            return JsonResponse({"errors": errors}, status=400)

        changed_fields = list(updates.keys())
        if updates:
            for field, value in updates.items():
                setattr(banner, field, value)
            banner.save(update_fields=changed_fields + ["updated_at"])

        record_admin_action(
            admin=request.current_user,
            action_type="banner.update",
            target_type="banner",
            target_id=banner.id,
            metadata={"changed_fields": changed_fields},
        )
        banner.refresh_from_db()
        return JsonResponse(_serialize_banner(banner))

    return JsonResponse({"error": "Method not allowed"}, status=405)
